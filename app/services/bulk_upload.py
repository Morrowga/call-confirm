"""Bulk CSV/Excel upload for appointments and event fan lists.

Pipeline: parse file -> lightweight AI model maps inconsistent columns to the
required fields -> hard validation layer (phone format, duplicates, date/time
sanity with timezone correctness, past-date rejection) -> summary.

The AI step only proposes a {source_column: required_field} mapping; it never
bypasses validation. If no AI key is configured, a heuristic mapper is used.
"""
import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from app.core.config import settings
from app.models import ConfirmationType
from app.services import confirmation_templates as templates
from app.services import csv_rephrase

APPOINTMENT_FIELDS = ["client_name", "phone_number", "date", "time", "timezone", "language"]
FAN_FIELDS = ["name", "phone_number"]

_HEURISTICS = {
    "client_name": ["client", "customer", "patient", "name", "full name"],
    "name": ["name", "fan", "contact", "full name"],
    "phone_number": ["phone", "mobile", "cell", "number", "tel", "msisdn"],
    "date": ["date", "day", "appointment date"],
    "time": ["time", "hour", "appointment time"],
    "timezone": ["timezone", "tz", "time zone"],
    "language": ["language", "lang", "locale"],
}


def heuristic_map(headers: list[str], required: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for target in required:
        for header in headers:
            h = header.strip().lower()
            if any(alias in h for alias in _HEURISTICS.get(target, [target])):
                if header not in mapping:
                    mapping[header] = target
                    break
    return mapping


def ai_map_columns(headers: list[str], sample_rows: list[dict], required: list[str]) -> dict[str, str]:
    """Use a lightweight, low-cost model (gpt-5-mini) to map columns.
    Falls back to heuristics when no API key is configured or the call fails."""
    api_key = settings.openai_api_key
    if not api_key:
        return heuristic_map(headers, required)
    try:
        import openai
        client = openai.OpenAI(api_key=api_key)
        prompt = (
            "Map the source spreadsheet columns to the required fields. "
            f"Source columns: {headers}. Sample rows: {sample_rows[:3]}. "
            f"Required fields: {required}. "
            'Respond ONLY with JSON: {"<source column>": "<required field>"} '
            "including only confident mappings."
        )
        completion = client.chat.completions.create(
            model="gpt-5-mini",
            max_completion_tokens=1200,
            reasoning_effort="low",
            response_format={"type": "json_object"},
            messages=[{"role": "user", "content": prompt}],
        )
        text = completion.choices[0].message.content.strip()
        mapping = json.loads(text)
        return {k: v for k, v in mapping.items() if k in headers and v in required}
    except Exception:
        return heuristic_map(headers, required)


PHONE_RE = re.compile(r"^\+?[1-9]\d{6,14}$")


def valid_phone(raw: str) -> str | None:
    cleaned = re.sub(r"[\s\-().]", "", raw or "")
    return cleaned if PHONE_RE.match(cleaned) else None


# --- Strict-format CSV upload (appointments) --------------------------------
# Replaces AI-based flexible column mapping for this specific flow: the
# account picks ONE timezone and ONE language for the whole batch (not CSV
# columns), so every row must use exactly these headers, and reason/type
# validity is checked per-row. All-or-nothing: since uploads are capped at
# 100 rows, requiring a clean file rather than accepting partial success
# keeps the review step simple — nothing shows until everything is valid.
REQUIRED_APPOINTMENT_COLUMNS = [
    "name", "phone number", "confirmation type", "reason", "scheduled for", "schedule to send", "language",
]
REQUIRED_FAN_COLUMNS = ["name", "phone number"]
MAX_CSV_ROWS = 100
EMPTY_PLACEHOLDER = "-"

_CSV_DT_FORMATS = [
    "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M",
    "%d-%m-%Y %H:%M", "%Y/%m/%d %H:%M", "%d/%m/%Y %I:%M %p", "%m/%d/%Y %I:%M %p",
]


def _parse_csv_datetime(raw: str) -> datetime | None:
    raw = raw.strip()
    for fmt in _CSV_DT_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


@dataclass
class CsvValidationResult:
    success: bool
    errors: list[str] = field(default_factory=list)  # row-numbered, shown verbatim to the user
    rows: list[dict] = field(default_factory=list)   # only populated when success is True
    duplicates: int = 0  # skipped silently, not treated as a blocking error


def validate_appointments_csv(
    filename: str, content: bytes, timezone: str,
    available_types: list[str], valid_languages: set[str],
) -> CsvValidationResult:
    """`timezone` is the one global setting for the whole batch; `language`
    is per-row (a CSV column), since a single business can genuinely have
    customers who speak different languages."""
    try:
        headers, raw_rows = parse_file(filename, content)
    except Exception:
        return CsvValidationResult(success=False, errors=["Could not read the file — check it's a valid CSV or Excel file."])

    normalized_headers = {h.strip().lower() for h in headers}
    missing = [c for c in REQUIRED_APPOINTMENT_COLUMNS if c not in normalized_headers]
    if missing:
        return CsvValidationResult(
            success=False,
            errors=[f"Missing required column(s): {', '.join(missing)}. Download the sample file for the exact format."],
        )

    if len(raw_rows) == 0:
        return CsvValidationResult(success=False, errors=["The file has no data rows."])
    if len(raw_rows) > MAX_CSV_ROWS:
        return CsvValidationResult(
            success=False, errors=[f"The file has {len(raw_rows)} rows — the maximum allowed is {MAX_CSV_ROWS}."],
        )

    errors: list[str] = []
    rows: list[dict] = []

    for i, raw in enumerate(raw_rows, start=2):  # row 1 is the header
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
        row_ok = True

        name = row.get("name", "")
        if not name:
            errors.append(f"Row {i}: name is required")
            row_ok = False

        phone = valid_phone(row.get("phone number", ""))
        if not phone:
            errors.append(f"Row {i}: invalid or missing phone number")
            row_ok = False

        conf_type = row.get("confirmation type", "")
        if conf_type not in available_types:
            errors.append(f"Row {i}: '{conf_type}' is not a valid confirmation type for your account")
            row_ok = False

        language = row.get("language", "")
        if language not in valid_languages:
            errors.append(f"Row {i}: '{language}' is not a supported call language")
            row_ok = False

        reason = row.get("reason", "")
        if not reason:
            errors.append(f"Row {i}: reason is required")
            row_ok = False

        scheduled_for_raw = row.get("scheduled for", "")
        schedule_to_send_raw = row.get("schedule to send", "")
        scheduled_at = None
        conf_type_enum = ConfirmationType(conf_type) if row_ok else None

        if row_ok:  # only worth checking the time columns once type is known-valid
            needs_time = templates.has_scheduled_time(conf_type_enum)
            if needs_time:
                if not scheduled_for_raw or scheduled_for_raw == EMPTY_PLACEHOLDER:
                    errors.append(f"Row {i}: 'scheduled for' is required for '{conf_type}'")
                    row_ok = False
                else:
                    scheduled_at = _parse_csv_datetime(scheduled_for_raw)
                    if scheduled_at is None:
                        errors.append(f"Row {i}: unparseable 'scheduled for' value")
                        row_ok = False
            else:
                if not schedule_to_send_raw or schedule_to_send_raw == EMPTY_PLACEHOLDER:
                    errors.append(f"Row {i}: 'schedule to send' is required for '{conf_type}'")
                    row_ok = False
                else:
                    scheduled_at = _parse_csv_datetime(schedule_to_send_raw)
                    if scheduled_at is None:
                        errors.append(f"Row {i}: unparseable 'schedule to send' value")
                        row_ok = False

        if not row_ok:
            continue

        # AI rephrasing (gpt-5-mini) — fits the user's raw text correctly
        # into the template's slot (e.g. removes a redundant "for" the user
        # typed that duplicates the template's own wording) and identifies
        # proper nouns for SSML pronunciation at call time. Always produces
        # a best-effort result in the row's own declared language; a row is
        # never rejected over phrasing.
        hint = templates.template_hint(conf_type_enum, language)
        result = csv_rephrase.rephrase_reason(reason, language, hint)

        rows.append({
            "client_name": name,
            "phone_number": phone,
            "scheduled_at": scheduled_at,
            "timezone": timezone,
            "language": language,
            "confirmation_type": conf_type,
            "subject_detail": result["corrected"],
            "subject_detail_proper_nouns": result["proper_nouns"],
        })

    if errors:
        return CsvValidationResult(success=False, errors=errors)
    return CsvValidationResult(success=True, rows=rows)


def sample_appointments_csv() -> bytes:
    """The downloadable sample file shown in the upload tab — exact required
    format, with example rows showing both a scheduled-time type and a
    send-time-only type, and different languages per row (since language is
    per-row, not global — a batch can genuinely mix customers who speak
    different languages)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(REQUIRED_APPOINTMENT_COLUMNS)
    writer.writerow([
        "Jane Doe", "+15551234567", "appointment", "your annual checkup",
        "2026-09-01 15:00", EMPTY_PLACEHOLDER, "en",
    ])
    writer.writerow([
        "田中太郎", "+819012345678", "order", "iPhone 17 Pro, 256GB",
        EMPTY_PLACEHOLDER, "2026-09-01 09:00", "ja",
    ])
    return buf.getvalue().encode("utf-8")


def validate_fans_csv(filename: str, content: bytes, existing_phones: set[str] | None = None) -> CsvValidationResult:
    """Strict fixed-column format, same discipline as
    validate_appointments_csv — replaces the older AI-column-mapping
    approach (process_fans/ai_map_columns below), which didn't validate
    against a known format the way the appointments CSV flow does. `name`
    is optional here (FanContact.name is nullable) — only `phone number`
    is actually required.

    `existing_phones` — phone numbers already saved for this event (from a
    prior upload or public signups) — is checked in addition to duplicates
    within this file itself, so re-uploading the same list (or a list that
    overlaps with people who already signed up publicly) doesn't create
    duplicate contacts."""
    try:
        headers, raw_rows = parse_file(filename, content)
    except Exception:
        return CsvValidationResult(success=False, errors=["Could not read the file — check it's a valid CSV or Excel file."])

    normalized_headers = {h.strip().lower() for h in headers}
    missing = [c for c in REQUIRED_FAN_COLUMNS if c not in normalized_headers]
    if missing:
        return CsvValidationResult(
            success=False,
            errors=[f"Missing required column(s): {', '.join(missing)}. Download the sample file for the exact format."],
        )

    if len(raw_rows) == 0:
        return CsvValidationResult(success=False, errors=["The file has no data rows."])
    if len(raw_rows) > MAX_CSV_ROWS:
        return CsvValidationResult(
            success=False, errors=[f"The file has {len(raw_rows)} rows — the maximum allowed is {MAX_CSV_ROWS}."],
        )

    errors: list[str] = []
    rows: list[dict] = []
    seen: set[str] = set(existing_phones or set())
    duplicate_count = 0

    for i, raw in enumerate(raw_rows, start=2):  # row 1 is the header
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}

        phone = valid_phone(row.get("phone number", ""))
        if not phone:
            errors.append(f"Row {i}: invalid or missing phone number")
            continue
        if phone in seen:
            # Not a blocking error — duplicates are expected (re-uploading
            # a list, or overlap with public signups) and are just skipped,
            # reported as a count rather than a per-row message.
            duplicate_count += 1
            continue
        seen.add(phone)
        rows.append({"name": row.get("name") or None, "phone_number": phone})

    if errors:
        return CsvValidationResult(success=False, errors=errors)
    return CsvValidationResult(success=True, rows=rows, duplicates=duplicate_count)


def sample_fans_csv() -> bytes:
    """The downloadable sample file shown in the event contact-upload UI —
    exact required format for self-explanatory bulk contact uploads."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(REQUIRED_FAN_COLUMNS)
    writer.writerow(["Jane Doe", "+15551234567"])
    writer.writerow(["田中太郎", "+819012345678"])
    writer.writerow(["", "+84329874755"])  # name is optional
    return buf.getvalue().encode("utf-8")


@dataclass
class UploadSummary:
    total: int = 0
    valid: int = 0
    duplicates: int = 0
    invalid: int = 0
    errors: list[str] = field(default_factory=list)
    rows: list[dict] = field(default_factory=list)
    mapping: dict = field(default_factory=dict)

    @property
    def message(self) -> str:
        return (
            f"{self.total} uploaded — {self.valid} valid, "
            f"{self.duplicates} duplicates, {self.invalid} invalid"
        )


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    if filename.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip() for h in rows[0]]
        data = [dict(zip(headers, (str(v or "").strip() for v in r))) for r in rows[1:]]
        return headers, data
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    return list(headers), [dict(r) for r in reader]


def process_appointments(filename: str, content: bytes, default_tz: str) -> UploadSummary:
    headers, raw_rows = parse_file(filename, content)
    mapping = ai_map_columns(headers, raw_rows, APPOINTMENT_FIELDS)
    summary = UploadSummary(total=len(raw_rows), mapping=mapping)
    seen: set[tuple[str, str]] = set()
    now = datetime.now(ZoneInfo("UTC"))

    for i, raw in enumerate(raw_rows, start=2):
        row = {mapping[k]: v for k, v in raw.items() if k in mapping}
        phone = valid_phone(row.get("phone_number", ""))
        if not phone:
            summary.invalid += 1
            summary.errors.append(f"row {i}: invalid phone number")
            continue
        tz_name = row.get("timezone") or default_tz
        try:
            tz = ZoneInfo(tz_name)
        except ZoneInfoNotFoundError:
            summary.invalid += 1
            summary.errors.append(f"row {i}: unknown timezone '{tz_name}'")
            continue
        dt = _parse_datetime(row.get("date", ""), row.get("time", ""))
        if dt is None:
            summary.invalid += 1
            summary.errors.append(f"row {i}: unparseable date/time")
            continue
        dt = dt.replace(tzinfo=tz)
        if dt <= now:
            summary.invalid += 1
            summary.errors.append(f"row {i}: appointment is in the past")
            continue
        key = (phone, dt.isoformat())
        if key in seen:
            summary.duplicates += 1
            continue
        seen.add(key)
        summary.valid += 1
        summary.rows.append({
            "client_name": row.get("client_name", "Client"),
            "phone_number": phone,
            "scheduled_at": dt,
            "timezone": tz_name,
            "language": row.get("language") or "en",
        })
    return summary


def process_fans(filename: str, content: bytes) -> UploadSummary:
    headers, raw_rows = parse_file(filename, content)
    mapping = ai_map_columns(headers, raw_rows, FAN_FIELDS)
    summary = UploadSummary(total=len(raw_rows), mapping=mapping)
    seen: set[str] = set()
    for i, raw in enumerate(raw_rows, start=2):
        row = {mapping[k]: v for k, v in raw.items() if k in mapping}
        phone = valid_phone(row.get("phone_number", ""))
        if not phone:
            summary.invalid += 1
            summary.errors.append(f"row {i}: invalid phone number")
            continue
        if phone in seen:
            summary.duplicates += 1
            continue
        seen.add(phone)
        summary.valid += 1
        summary.rows.append({"name": row.get("name") or None, "phone_number": phone})
    return summary


_DT_FORMATS = [
    "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M",
    "%d-%m-%Y %H:%M", "%Y/%m/%d %H:%M", "%d/%m/%Y %I:%M %p", "%m/%d/%Y %I:%M %p",
]


def _parse_datetime(date_str: str, time_str: str) -> datetime | None:
    combined = f"{date_str.strip()} {time_str.strip()}".strip()
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(combined, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(combined)
    except ValueError:
        return None